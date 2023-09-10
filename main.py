import PySimpleGUI as sg
import os
from typing import List
import openpyxl
from openpyxl.utils import get_column_letter
import re

# In[150]:


sg.theme("LightGrey1")


# ## Functions for power

# In[154]:


# считывание
def el_read(element: List[str], path_to_file=None):
    if (element[0][0] == 'q') or (element[0][0] == 'Q'):
        if element[1] == '0':
            return '.meas ' + element[0] + '_power ABS AVG Ic(' + element[0] + ')*(-V(' + element[3] + ')' + '\n'
        elif element[3] == '0':
            return '.meas ' + element[0] + '_power ABS AVG Ic(' + element[0] + ')*V(' + element[1] + ')' + '\n'
        return '.meas ' + element[0] + '_power ABS AVG Ic(' + element[0] + ')*V(' + element[1] + ',' + element[3] + \
               ')' + '\n'
    elif (element[0][0] == 'r') or (element[0][0] == 'R'):
        if element[1] == '0':
            return '.meas ' + element[0] + '_power ABS AVG I(' + element[0] + ')*(-V(' + element[2] + '))' + '\n'
        elif element[2] == '0':
            return '.meas ' + element[0] + '_power ABS AVG I(' + element[0] + ')*V(' + element[1] + ')' + '\n'
        return '.meas ' + element[0] + '_power ABS AVG I(' + element[0] + ')*V(' + element[1] + ',' + element[2] + \
               ')' + '\n'
    elif (element[0] == 'd') or (element[0][0] == 'D'):
        rs = rs_giver(element[3], path_to_file, element[0])
        return '.meas ' + element[0] + '_power ABS AVG I(' + element[0] + ')*I(' + element[0] + ')*' + rs + '\n'
    elif (element[0] == 'm') or (element[0][0] == 'M'):
        if element[1] == '0':
            return '.meas ' + element[0] + '_power ABS AVG Id(' + element[0] + ')*(-V(' + element[3] + ')' + '\n'
        elif element[3] == '0':
            return '.meas ' + element[0] + '_power ABS AVG Id(' + element[0] + ')*V(' + element[1] + ')' + '\n'
        return '.meas ' + element[0] + '_power ABS AVG Id(' + element[0] + ')*V(' + element[1] + ',' + element[3] + \
               ')' + '\n'
    elif ((element[0][0] == 'X') and (element[0][1] == 'U')):  # Ix(U1:D)V(vdrain)
        minuselement = element[0]
        if element[1] == '0':
            return '.meas ' + minuselement[1:] + '_power ABS AVG Ix(' + minuselement[1:] + ':D)*(-V(' + element[
                3] + ')' + '\n'
        elif element[3] == '0':
            return '.meas ' + minuselement[1:] + '_power ABS AVG Ix(' + minuselement[1:] + ':D)*V(' + element[
                1] + ')' + '\n'
        return '.meas ' + minuselement[1:] + '_power ABS AVG Ix(' + minuselement[1:] + ')*V(' + element[1] + ',' + \
               element[3] + \
               ')' + '\n'


# In[155]:
def choose_formula(element, r):
    layout = [
        [sg.Text('Выберите формулу для расчета элемента ' + element[0])],
        [sg.Radio('RMS', 'Radio1', default=False, key='-IN1')],
        [sg.Radio('AVG', 'Radio1', default=True, key='-IN2')],
        [sg.Button('Ввод')],
    ]

    window = sg.window = sg.Window('Выбор формулы', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        # if event in (sg.WINDOW_CLOSED,'Cancel'):
        #     value = ''
        #     break
        if event == 'Ввод':
            if values['-IN1'] == True:
                value = '.MEASURE TRAN ' + element[0] + '_power RMS (V(' + element[1] + ', ' + element[
                    3] + ')*I(' + r + '))\n'

            elif values['-IN2'] == True:
                value = '.MEASURE TRAN ' + element[0] + '_power AVG (ABS(V(' + element[1] + ', ' + element[
                    3] + '))*ABS(I(' + r + ')))\n'

        window.close()
        return value, True


def el_read_ckt(element: List[str], r=None, path_to_file=None):
    if (element[0][0] == 'q') or (element[0][0] == 'Q'):
        if r != None:
            return '.MEASURE TRAN ' + element[0] + '_power RMS(V(' + element[1] + ', ' + element[
                3] + ')*I(' + r + '))\n', True
        else:
            sg.Popup('Подключите сопротивление к транзистору ' + element[0] + '!', title=errt, button_type=0,
                     icon='warning.ico')
            return '', False
    elif (element[0] == 'm') or (element[0][0] == 'M'):
        if r != None:
            return choose_formula(element,
                                  r)  # '.MEASURE TRAN ' + element[0] + '_power RMS(V(' + element[1] + ', ' + element[3] + ')*I(' + r + '))\n', True
        else:
            sg.Popup('Подключите сопротивление к транзистору ' + element[0] + '!', title=errt, button_type=0,
                     icon='warning.ico')
            return '', False
    elif (element[0][0] == 'r') or (element[0][0] == 'R'):
        return '.MEASURE TRAN ' + element[0] + '_power RMS I(' + element[0] + ')*V(' + element[2] + ',' + element[1] + \
               ')' + '\n', True
    elif (element[0] == 'd') or (element[0][0] == 'D'):
        return '.MEASURE TRAN ' + element[0] + '_power AVG ABS(V(' + element[0] + '))*ABS(I(' + element[
            0] + '))\n', True
    elif (element[0][0] == 'X'):
        if r != None:
            return choose_formula(element,
                                  r)  # '.MEASURE TRAN ' + element[0] + '_power RMS(V(' + element[1] + ', ' + element[3] + ')*I(' + r + '))\n', True
        else:
            sg.Popup('Подключите сопротивление к транзистору ' + element[0] + '!', title=errt, button_type=0,
                     icon='warning.ico')
            return '', False


# In[155]:


# функция перевода файла в текстовый массив, если файл не в формате .txt
def file2strarr_enc(name):
    file = None
    try:
        with open(name, encoding='utf-16-le') as fh:  #
            file = fh.read()
            mylix = file.splitlines(True)
            return mylix
    except FileNotFoundError:
        sg.Popup('Файл ' + name + ' не найден.', title=errt, button_type=0, icon='warning.ico')
        return 'None'


# In[156]:


# функция перевода файла в текстовый массив
def file2strarr(name):
    try:
        sor = open(name, 'r')
        mylix = sor.read().splitlines(True)
        sor.close()
        return mylix
    except FileNotFoundError:
        sg.Popup('Файл ' + name + ' не найден.', title=errt, button_type=0, icon='warning.ico')
        return 'None'

    # ## Variables


# In[157]:


# ввод глобальных элементов для списков элементов, адреса и сохранения выбранных элементов
k = []
adrs = []
comp_names = ['R1', 'C1', 'R2', 'C2', 'R3', 'C3', 'R4', 'C4', 'R5', 'C5']
slist = []
nods = ['0', 'Tj_', 'T1_', 'T2_', 'T3_', 'T4_', 'T5_', 'Ts_', 'Tamb']
new_net_w_temp = 0

# In[158]:


# здесь находятся тексотовые переменные.
text_exit_button = 'Выйти'
text_browse_file = "1. Выберите файл-netlist вашей схемы в формате .txt, .net, .cir или .CKT:"
text_browse_excel = "2. Выберите файл excel с тепловыми параметрами компонентов:   "
text_open, key_open, key_open_net = "Выбрать файл", "-IN-", "-myinput-"
key_r1, key_r2, key_r3, key_r4, key_r5, key_ramb = "-myinputr1-", "-myinputr2-", "-myinputr3-", "-myinputr4-", "-myinputr5-", "-myinputr6-"
key_c1, key_c2, key_c3, key_c4, key_c5 = "-myinputc1-", "-myinputc2-", "-myinputc3-", "-myinputc4-", "-myinputc5-"
text_show_the_list = "Вывести список элементов"
text_choose_elements, key_list, key_list_rad = "3. Выберите элементы для создания тепловых моделей \n(среди биполярных и МОП-транзисторов, \n\
резисторов и диодов): ", "-listbox-", "-listbox2-"
text_choose_rads = "Выберите подходящий радиатор для "
text_help = '4. Для выбора элемента щёлкните по нему левой кнопкой мыши,для выбора нескольких элементов зажмите клавишу \
Ctrl и щёлкните по нужным элементам левой кнопкой мыши. \nПосле выбора элементов нажмите "Сохранить", чтобы записать \
команды для расчёта мощностей выбранных компонентов.'
text_save = "Сохранить"
text_copy = "Копировать"
text_close = "Закрыть"
text_for_error_file_open = '5. После запуска файла со схемой в симуляторе нажмите "Сделать тепловую модель" для обработки файла с выходными данными, чтобы создать Netlist файл в формате .cir или .CKT.'
text_power = "Мощность выбранных компонентов:\n"
text_rs_error = 'Ошибка: не посчитана мощность\nТакого диода нет в базе'
text_enter_param_1 = 'Введите тепловые параметры компонента '
text_enter_param_2 = '\nR1 - тепловое сопротивление от корпуса до среды; С1 - теплоемкость'
text_enter_param_caution = 'Если параметров меньше пяти, то в оставшиеся поля впишите 0\nНажмите Cancell, чтобы пропустить и не включать компонент в netlist'
text_enter_pulsentime = 'Введите параметры источника и время симуляции'
text_enter_time = 'Введите время симуляции '
text_enter_board = 'Введите размеры платы'
text_choice = '     Выберите подходящий вариант     '
text_ask_temp_change = "Пересчитать схему с учетом полученных температур?\nПримечание: для этого в следующем окне нужно открыть полученный netlist."
text_temp_change_spice = 'Для того, чтобы пересчитать схему с учетом температур:\n1) Проведите симуляцию созданного netlist\'a\n2) Нажмите \"Записать температуры\".\n3) Netlist схемы создан с учетом температур. Проведите расчет схемы и выполните программу с пункта 5. \
\n\nПримечание: для того, чтобы стабилизировать температуру на БТ, поставьте команду TEMP={T_Х} на подключенный к нему диод, если такой имеется. Вместо Х напишите обозначение этого транзистора.'
title_temp_change = 'Учет температуры'
text_temp_change_microcap = 'Для того, чтобы пересчитать схему с учетом температур:\n1) Проведите симуляцию созданного netlist\'a\n2) После симуляции нажмите F12 (State Variables), в появившемся окне нажмите Print. \
\n3) Нажмите \"Записать температуры\"\n4) Netlist схемы создан с учетом температур. Проведите расчет схемы и выполните программу с пункта 5. \
\n\nПримечание: для того, чтобы стабилизировать температуру на БТ, поставьте команду T_ABS={T_Х} на подключенный к нему диод, если такой имеется. Вместо Х напишите обозначение этого транзистора.'
errt = "Ошибка"
chft = 'ᅠ ᅠ ᅠ ᅠ ᅠ ᅠ Выберите netlist-файл.ᅠ ᅠ ᅠ ᅠ ᅠ ᅠ '
chft2 = 'ᅠ ᅠ ᅠ ᅠ ᅠ ᅠ Выберите файл Базы Данных. ᅠ ᅠ ᅠ ᅠ ᅠ '
ches = 'ᅠ ᅠ ᅠ ᅠ ᅠ ᅠ Выберите элементы!ᅠ ᅠ ᅠ ᅠ ᅠ ᅠ '
text_rad = '    Поставить радиатор на элемент '
title_rad = "Радиатор"
errt2 = 'Пустой файл'
text1_open_error = 'Открыть файл ошибки в другой папке'
key_open_error = '-errorfile-'
text_open_error_default = 'Сделать тепловую модель'
open_net_file = 'Открыть полученный файл'
error_text_error_log = 'ᅠ ᅠ ᅠ ᅠ ᅠ ᅠ Файл error_log отсутствует!ᅠ ᅠ ᅠ ᅠ ᅠ ᅠ '
final_popup_text = 'Программа создала тепловую модель. Нажмите \"Открыть полученный файл\" для просмотра.'
text_hint = 'Для работы с программой необходим Netlist файл схемы. Чтобы получить файл Netlist:\n\n В LTSpice: Откройте схему и в верхнем меню View выберите \"SPICE Netlist\". \
В папке со схемой появится временный файл с форматом данных .net\n(Можно сохранить файл в формате .cir, щелкнув ПКМ\
 по диалоговому окну Netlist,чтобы сделать файл постоянным, и иметь возможность работать с ним вне\
 зависимости от LTSpice).\n\n \
В Micro Cap: Откройте схему и в верхнем меню File выберите Translate -> Schematic to SPICE text file.\nВ появившемся \
диалоговом окне "Translate to SPICE" в графе Spice type выбрать Micro Cap spice, в графе Options выбрать .options и Expand SUBCKT. \n\
Выбрать путь, куда сохранится файл в формате .CKT и нажать ОК. ВАЖНО! Для рассчета мощности транзисторов к стоку (для МОПТ)/ к коллектору (для БТ) должно быть подключено сопротивление!'
message_treatment = 'Чтобы открыть Netlist файл нажмите на кнопку "Открыть файл". Открыв файл запустите его симуляцию.\n\
В LTspice: Simulate → Run.\n\
В Micro Cap: Analisys → Transient. В открывшейся вкладке, до того как нажать Run, нажмите Ctrl+Alt+M. В новом окне нажмите \
на зеленый плюс и из выпадающего списка Function выберите функции рассчета, которые записала программа (напр. M1_POWER).\n\
После этого можно запускать симуляцию, нажав Run.'
rus_lower = set('аоуыэяеёюибвгдйжзклмнпрстфхцчшщ')
rus_upper = set('АОУЫЭЯЕЁЮИБВГДЙЖЗКЛМНПРСТФХЦЧШЩ')
text_for_somebody = 'Вы можете выйти из программы, нажав кнопку "Выйти". \nВы также можете выполнить расчет заново \
или выполнить его для новой схемы, для этого начните выполнять действия с первого шага.'
blunt_text = '                                                         \n'
ckt_measure_text_start = '[Grid Text]\nText="'
ckt_measure_text_end = '"\nPx=16,2040\nGridSnap=True\nJustifyH=Left\nJustifyV=Bottom\nCBorder=None'


# ## Functions for reader_saver

# In[159]:


def mes_generator(svadrs: str, inmes: str):
    mes = 'Откройте файл со схемой, Netlist которой обрабатывали.\nДиректория Netlist — ' + svadrs + '\nИ соответствует\
        директории схемы.\n' + 'Для дальнейших расчетов вам необходимо добавить на схему следующую команду: ' + '\n' + \
          inmes + '\nЧтобы скопировать команду в буфер обмена, нажмите "Копировать"'
    return mes


# In[160]:


def include_address(ss: str, a):
    down = ''
    for i in ss:
        if i != '.':
            down += i
        elif i == '.':
            break
    if a == 1:
        down += '-1.inc'
    elif a == 2:
        down += '-2.inc'
    return down


# In[161]:


def error_file_adrs(a: str, net_format: str):
    if net_format == 'cir':
        down = a.replace('.cir', '.log')
        return down
    if net_format == 'ckt':
        down = a.replace('.CKT', '.TNO')
        return down


# In[161]:


def final_file_adrs_creator(svadrs, name):
    final_file = svadrs + '/тепловая модель ' + name
    final_file_w_temp = svadrs + '/тепловая модель с учетом температур ' + name
    return final_file, final_file_w_temp


# In[162]:


def listreader(a: str):
    mylix = file2strarr(a)
    if ('\x00' in mylix[1]) or ('\x00' in mylix[0]):
        mylix = file2strarr_enc(a)
    if mylix != 'None':
        schema_adrs = mylix[0].replace('* ', '')
        schema_adrs = schema_adrs.replace('\n', '')
        names = []
        list_with_info = []
        for line in mylix:
            gar = line.split()
            if gar != []:
                if (gar[0][0] == 'Q') or (gar[0][0] == 'R') or (gar[0][0] == 'M') or (gar[0][0] == 'D') or (
                        (gar[0][0] == 'X') and (gar[0][1] == 'U')) or (gar[0][0] == 'X'):
                    names.append(gar[0])
                    list_with_info.append(line)
        for n in range(len(list_with_info)):
            if 'PARAMS' in list_with_info[n]:
                list_with_info[n] = list_with_info[n].split(' ')
                list_with_info[n].pop().pop()
                list_with_info[n] = ' '.join(list_with_info[n])
        return mylix, names, schema_adrs, list_with_info
    else:
        return None, None, None, None


# In[163]:

# jnn
def saving(a1: List[str], a2: str, net_format: str, z=None, path_to_xsl=None):
    if net_format == 'cir':
        if path_to_xsl != None:
            f = open(a2, 'r')
            mylix = file2strarr(a2)
            if ('\x00' in mylix[1]) or ('\x00' in mylix[0]):
                f.close()
                f = open(a2, 'r', encoding='utf-16-le')
            functions = ''
            netlist_text = ''
            line = f.readline()
            while line:
                if 'TEMP' not in line:
                    netlist_text += line
                if line[0] == '.' and line[1] == 'b':
                    for iz in a1:
                        se = iz.split()
                        prov = se[0]
                        print(prov)
                        for stroka in z:
                            if stroka == prov:
                                functions = functions + (el_read(se, path_to_xsl))
                    netlist_text += functions + '.tran 1m\n.end\n\n'
                    f.close()
                    f = open(a2, 'w', encoding='utf-16-le')
                    f.write(netlist_text)
                    f.close()
                    return True
                    break
                line = f.readline()
        else:
            f = open(a2, 'w', encoding='utf-16-le')
            for a in a1:
                f.write(a)
    else:
        if path_to_xsl != None:
            netlist_text = ''
            functions = ''
            f = open(a2, 'r')
            line = f.readline()
            while line:
                if 'T_ABS' not in line:
                    if 'PARAMS' not in line:
                        netlist_text += line
                if ('.END' in line) and ('S' not in line):  # [0] == '.' and line[1] == 'E'
                    s = netlist_text.split('\n')
                    while True:
                        if s[-1] != '*':
                            s.pop()
                        else:
                            break
                    netlist_text = '\n'.join(s)
                    for iz in a1:
                        se = iz.split()
                        prov = se[0]
                        for stroka in z:
                            if stroka == prov:
                                if ('M' in stroka) or ('Q' in stroka) or ('X' in stroka):
                                    R = r_getter(se, a1)
                                    text, check = el_read_ckt(se, R)
                                    functions = functions + text
                                else:
                                    text, check = el_read_ckt(se, path_to_xsl)
                                    functions = functions + text
                    netlist_text += '\n' + functions + '.END'
                    f.close()
                    f = open(a2, 'w')
                    f.write(netlist_text)
                    f.close()
                    return check
                    break
                line = f.readline()
        else:
            f = open(a2, 'w')
            for a in a1:
                f.write(a)


# In[163.5]:

def saving_temp(a1: List[str], a2: str, net_format: str, include_adrs: str, z):
    if net_format == 'cir':
        f = open(a2, 'r', encoding='utf-16-le')
        netlist_text = ''
        line = f.readline()
        while line:
            netlist_text += line
            for iz in a1:
                se = iz.split()
                if iz == line.replace('\n', ''):
                    netlist_text += '+ TEMP={T_' + se[0] + '}\n'
            if line[0] == '.' and line[1] == 'e':
                netlist_text = netlist_text.split('\n')
                netlist_text.pop()
                netlist_text.pop()
                netlist_text = '\n'.join(netlist_text)
                netlist_text += '\n.include ' + include_adrs.replace('/', '\\') + '\n.end'
            line = f.readline()
        f.close()
        f = open(a2, 'w', encoding='utf-16-le')
        f.write(netlist_text)
        f.close()
    else:
        netlist_text = ''
        f = open(a2, 'r')
        line = f.readline()
        check = False
        x_elem = False
        while line:
            if 'T_ABS' not in line:
                netlist_text += line
            for iz in a1:
                se = iz.split()
                # if line[0] == 'X':
                if iz == line.replace('\n', ''):
                    if line[0] == 'X':
                        netlist_text += '+ PARAMS: TNOM={T_' + se[0] + '}\n'
                        x_elem = True
                        break
                    else:
                        break
                chk_upper = ('.SUBCKT ' + se[-1]) in line
                chk_lower = ('.SUBCKT ' + se[-1].lower()) in line
                if (chk_upper and x_elem) or (chk_lower and x_elem):
                    netlist_text += '+ PARAMS: TNOM=27\n'
                    check = True
                    # x_elem = False
                    break
                if check and ('.MODEL' in line) and (('NMOS' in line) or ('PMOS' in line)):
                    check = False
                    netlist_text += '+ T_ABS={TNOM}\n'
                    break
                if ('.MODEL ' + se[-1]) in line:
                    netlist_text += '+ T_ABS={T_' + se[0] + '}\n'
                    break
            if line == '.END':
                netlist_text = netlist_text.split('\n')
                netlist_text.pop()
                netlist_text = '\n'.join(netlist_text)
                netlist_text += '\n.INCLUDE ' + include_adrs.replace('/', '\\') + '\n.END'
            line = f.readline()
        f.close()
        f = open(a2, 'w')
        f.write(netlist_text)
        f.close()


# In[163.5]:

def r_getter(se, a1):
    a = False
    for iz in a1:
        stroka = iz.split()
        if ('R' in stroka[0]) and ((se[1] == stroka[1]) or (se[1] == stroka[2])):
            a = True
            return stroka[0]
            break
    if not a:
        return None


# In[163.5]:

def rs_giver(name, path_to_file, ref):
    wb = openpyxl.load_workbook(path_to_file)
    sheets_list = wb.sheetnames
    sheet_active = wb[sheets_list[2]]
    row_max = sheet_active.max_row

    search_text = name
    row_min = 1
    data_from_rs = ''

    while row_min <= row_max:
        row_min = str(row_min)

        word_cell = 'B' + row_min

        data_from_cell = str(sheet_active[word_cell].value)
        regular = search_text

        if regular == data_from_cell:
            rs_cell = 'O' + row_min
            data_from_rs = str(sheet_active[rs_cell].value)
            break

        row_min = int(row_min) + 1
    if data_from_rs != '':
        return data_from_rs
    else:
        rs_value = enter_self_rs(ref)[0]
        return str(rs_value)


# In[164]:
def enter_self_rs(ref):
    layout = [
        [sg.Text('Ввод паразитного сопротивления диода ' + ref)],
        [sg.Text('Rs'), sg.InputText(size=(5, 5), key=key_r1)],
        [sg.Button('OK'), sg.Button('Cancel')],
    ]
    window = sg.Window('Ввод параметров', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            value = 'None'
            break
        elif event == 'OK':
            if values[key_r1] == '':
                sg.Popup('Введите значение!', title=errt, button_type=0, icon='warning.ico')
            else:
                value = [values[key_r1]]
                break
    window.close()
    return value


# In[164]:


def read_with_decoder(name, first_name, n):
    print(name, 'with')
    file = None
    lix = []
    try:
        with open(name, encoding='utf-16-le') as fh:
            file = fh.read()
        lix = file.splitlines()
        # return(lix)
        print(lix)
        spisok = []
        for l in range(len(lix)):
            if lix[l].find(first_name) == 0:
                for i in range(l, l + n):
                    spisok.append(lix[i])
                break
        return spisok
    except FileNotFoundError:
        sg.Popup('Файл ' + name + ' не найден.', title=errt, button_type=0, icon='warning.ico')
        return 'None'


# In[165]:


def read_without_decoder(name, first_name, n):
    print(name, 'without')
    try:
        sor = open(name, 'r')
        mylix = sor.read().splitlines(True)
        print(mylix)
        sor.close()
        spisok = []
        for l in range(len(mylix)):
            if (first_name in mylix[l]) or (first_name.upper() in mylix[l]):
                for i in range(l, l + n):
                    spisok.append(mylix[i])
                break
        return spisok
    except FileNotFoundError:
        sg.Popup('Файл ' + name + ' не найден.', title=errt, button_type=0, icon='warning.ico')
        return 'None'


# In[166]:

# проверка на наличие символов кириллицы в названии или адресе
def checkk(a, first_name, n):
    # print(a)
    f = True
    rus_lower = 'аоуыэяеёюибвгдйжзклмнпрстфхцчшщ'
    rus_upper = 'АОУЫЭЯЕЁЮИБВГДЙЖЗКЛМНПРСТФХЦЧШЩ'
    for letter in a:
        if (rus_lower.find(letter) != -1) or (rus_upper.find(letter) != -1):
            f = False
            break
    if f == True:
        return read_without_decoder(a, first_name, n)
    else:
        return read_with_decoder(a, first_name, n)


# In[167]:

# сюда приходит адрес и выбранные элементы, на выходе формируется массив
def power_file_saving(address, first_name, number_of_elements, net_format, list_of_chosen_elements):
    n = number_of_elements
    spisok = checkk(address, first_name, n)
    print(spisok)
    power = []
    if net_format == 'cir':
        if spisok != 'None':
            for s in spisok:
                list1 = s.split('=')
                value = ''
                for letter in list1[1]:
                    if letter != ' ':
                        value += letter
                    else:
                        break
                power.append(str(float(value)))
            return power
        else:
            return []
    else:
        if (spisok != 'None') and (len(list_of_chosen_elements) == len(spisok)):
            for elem in list_of_chosen_elements:
                for s in spisok:
                    list1 = s.split()
                    if list1 != []:
                        if elem == list1[0].replace('_POWER', ''):
                            # value = list1[1]
                            power.append(str(list1[1]))
                            break
            return power
        else:
            sg.Popup('Не получилось достать мощности.', title=errt, button_type=0, icon='warning.ico')
            return []


# In[168]

def give_name(a1: List[str], net_format):
    if net_format == 'cir':
        if (stroka[0] == 'm') or (stroka[0][0] == 'M'):
            a = stroka[5]
        elif (stroka[0] == 'q') or (stroka[0][0] == 'Q'):
            a = stroka[5]
        elif (stroka[0] == 'd') or (stroka[0][0] == 'D'):
            a = stroka[3]
        elif ((stroka[0][0] == 'X') and (stroka[0][1] == 'U')):
            a = stroka[4]
        elif (stroka[0] == 'r') or (stroka[0][0] == 'R'):
            a = stroka[0]
        return a, ' '.join(stroka)
    else:
        if (stroka[0] == 'm') or (stroka[0][0] == 'M'):
            a = stroka[5]
        elif (stroka[0] == 'q') or (stroka[0][0] == 'Q'):
            a = stroka[4]
        elif (stroka[0] == 'd') or (stroka[0][0] == 'D'):
            a = stroka[3]
        elif ((stroka[0] == 'x') or (stroka[0][0] == 'X')):
            a = stroka[4]
        elif (stroka[0] == 'r') or (stroka[0][0] == 'R'):
            a = stroka[0]

        if '_' in a:
            a = a.replace("_", "")
        return a, ' '.join(stroka)


# In[169]
def replacing_lishnee(chosen_elems_info, element_names):
    for n in range(len(chosen_elems_info)):
        s = chosen_elems_info[n].split()
        while s[-1] != element_names[n]:
            s.pop()
        chosen_elems_info[n] = ' '.join(s)
    return chosen_elems_info


# In[169]
def rad_name_pick(rad_names, component_name):
    layout = [
        [sg.Text(text_choose_rads + component_name)],
        [sg.Listbox(rad_names, s=(60, 10), select_mode='extended', key=key_list_rad)],
        [sg.Button('OK'), sg.Button('Cancel')],
        [sg.Text()],
    ]
    window = sg.Window('Радиатор', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        window.find_element(key=key_list_rad).Update(rad_names)
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            value = 'None'
            break
        elif event == 'OK':
            value = values[key_list_rad]
        if event == 'OK' and value != [] and len(value) == 1:
            break
    window.close()
    return value


# In[169.5]

def self_enter(name):
    layout = [
        [sg.Text(name + ' - Такого компонента нет в базе')],
        [sg.Text(text_enter_param_1 + name)],
        [sg.Text(text_enter_param_caution)],
        [sg.Text('R1'), sg.InputText(size=(5, 5), key=key_r1), sg.Text('C1'), sg.InputText(size=(5, 5), key=key_c1)],
        [sg.Text('R2'), sg.InputText(size=(5, 5), key=key_r2), sg.Text('C2'), sg.InputText(size=(5, 5), key=key_c2)],
        [sg.Text('R3'), sg.InputText(size=(5, 5), key=key_r3), sg.Text('C3'), sg.InputText(size=(5, 5), key=key_c3)],
        [sg.Text('R4'), sg.InputText(size=(5, 5), key=key_r4), sg.Text('C4'), sg.InputText(size=(5, 5), key=key_c4)],
        [sg.Text('R5'), sg.InputText(size=(5, 5), key=key_r5), sg.Text('C5'), sg.InputText(size=(5, 5), key=key_c5)],
        [sg.Text('Ramb'), sg.InputText(size=(5, 5), key=key_ramb),
         sg.CBox('Радиатор', size=(10, 1), enable_events=True, key='rad')],
        [sg.Button('OK'), sg.Button('Cancel')],
    ]
    window = sg.Window('Ввод параметров', layout, finalize=True, modal=True)
    rad = False
    while True:
        event, values = window.read()

        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            value = 'None'
            rad = False
            ramb = 'None'
            break
        elif event == 'rad':
            rad = True
        elif event == 'OK':
            if values[key_r1] == '' and values[key_r2] == '' and values[key_r3] == '' and values[key_r4] == '' and \
                    values[
                        key_r5] == '' and values[key_ramb] == '' and values[key_c1] == '' and values[key_c2] == '' and \
                    values[key_c3] == '' and values[
                key_c4] == '' and values[key_c5] == '':
                sg.Popup('Заполните как минимум одну пару полей', title=errt, button_type=0, icon='warning.ico')
            else:
                value = [values[key_r1], values[key_c1], values[key_r2], values[key_c2], values[key_r3], values[key_c3],
                         values[key_r4], values[key_c4], values[key_r5], values[key_c5]]
                ramb = values[key_ramb]
                break
    window.close()
    return value, rad, ramb


# In[169.55]

def self_enter_rad():
    layout = [
        [sg.Text(text_enter_param_1 + ' - радиатора')],
        [sg.Text('R_th'), sg.InputText(size=(5, 5), key=key_r1), sg.Text('C_th'),
         sg.InputText(size=(5, 5), key=key_c1)],
        [sg.Button('OK'), sg.Button('Cancel')],
    ]
    window = sg.Window('Ввод параметров', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            value = 'None'
            break
        elif event == 'OK':
            if values[key_r1] == '' or values[key_c1] == '':
                sg.Popup('Заполните все поля', title=errt, button_type=0, icon='warning.ico')
            else:
                value = [values[key_r1], values[key_c1]]
                break
    window.close()
    return value


# In[169.55]

def self_enter_res(name):
    layout = [
        [sg.Text(text_enter_param_1 + name + text_enter_param_2)],
        [sg.Text('R1'), sg.InputText(size=(5, 5), key=key_r1), sg.Text('C1'),
         sg.InputText(size=(5, 5), key=key_c1)],
        [sg.Button('OK'), sg.Button('Cancel')],
    ]
    window = sg.Window('Ввод параметров', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            value = 'None'
            break
        elif event == 'OK':
            if values[key_r1] == '' or values[key_c1] == '':
                sg.Popup('Заполните все поля', title=errt, button_type=0, icon='warning.ico')
            else:
                value = [values[key_r1], values[key_c1]]
                break
    window.close()
    return value


# In[18]

def write_comp(f, comp_data_enter, elements, i, nod_counter, nc, Rjc):
    for j in range(len(comp_data_enter)):
        if comp_data_enter[j] != '0.0' and comp_data_enter[j] != '0' and comp_data_enter[j] != '':
            if j % 2 == 0:
                f.write(
                    comp_names[j] + '_' + elements[i] + ' ' + nods[nc + 1] + elements[i] + ' ' + nods[nc] + elements[
                        i] + ' ' + comp_data_enter[j] + '\n')
                Rjc += float(comp_data_enter[j])
                nc += 1

            if j % 2 != 0:
                f.write(comp_names[j] + '_' + elements[i] + ' ' + nods[nod_counter // 2] + elements[i] + ' 0 ' +
                        comp_data_enter[j] + '\n')

            nod_counter += 1
        else:
            break
    return Rjc


# In[180]
def enter_tran():
    layout = [
        [sg.Text(text_enter_time)],
        [sg.Text('.tran '), sg.InputText('1000', size=(8, 5), key=key_r1), ],
        [sg.Button('Ввести')],
        [sg.Text()], ]

    window = sg.Window('Ввод параметров симуляции', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event == 'Ввести':
            if values[key_r1] == []:
                sg.Popup('Заполните поле', title=errt, button_type=0, icon='warning.ico')
                break
            else:
                t = ' '.join([values[key_r1]])
                break

    window.close()
    return t


# In[18 ]

def choose_power(pulse_power, element_name):
    layout = [
        [sg.Text('Выберите тип режима источника для элемента ' + element_name)],
        [sg.Radio('none (при выборе будет задан постоянный источник V(power)', 'Radio1', default=False, key='-IN1')],
        [sg.Radio('PULSE', 'Radio1', default=True, key='-IN2')],
        [sg.Radio('SINE', 'Radio1', default=False, key='-IN3')],
        [sg.Radio('EXP', 'Radio1', default=False, key='-IN4')],
        [sg.Radio('SFFM', 'Radio1', default=False, key='-IN5')],
        [sg.Radio('PWL', 'Radio1', default=False, key='-IN6')],
        [sg.Button('Ввод'), sg.Button('Cancel')],
    ]

    window = sg.window = sg.Window('Выбор режима', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            value = 'None'
            break
        elif event == 'Ввод':
            if values['-IN1'] == True:
                value = pulse_power

            elif values['-IN2'] == True:
                value = enter_pulse(pulse_power, element_name)

            elif values['-IN3'] == True:
                value = power_sine(pulse_power, element_name)

            elif values['-IN4'] == True:
                value = power_exp(pulse_power, element_name)

            elif values['-IN5'] == True:
                value = power_sffm(pulse_power, element_name)

            elif values['-IN6'] == True:
                value = power_pwl(pulse_power, element_name)

            if value != 'None':
                break
    window.close()
    return value


# In[17 ]

def enter_pulse(pulse_power, element_name):
    layout = [
        [sg.Text(text_enter_pulsentime + ' для элемента ' + element_name)],
        [sg.Text('Vinitial[V]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r1), ],
        [sg.Text('Von[V]: заполняется автоматически '), ],
        [sg.Text('Tdelay[s]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r2), ],
        [sg.Text('Trise[s]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r3), ],
        [sg.Text('Tfall[s]: ', size=(8, 0)), sg.InputText('1', size=(8, 5), key=key_r4), ],
        [sg.Text('Ton[s]: ', size=(8, 0)), sg.InputText('1000', size=(8, 5), key=key_r5), ],
        [sg.Text('Tperiod[s]: ', size=(8, 0)), sg.InputText('1001', size=(8, 5), key=key_c1), ],
        [sg.Text('Ncycles: ', size=(8, 0)), sg.InputText('10000', size=(8, 5), key=key_c2), ],
        [sg.Button('Ввести'), sg.Button('Cancel')],
        [sg.Text()], ]

    window = sg.Window('Ввод параметров симуляции', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            res = 'None'
            break
        elif event == 'Ввести':
            if values[key_r1] == '' or values[key_r2] == '' or values[key_r3] == '' or values[key_r4] == '' or values[
                key_r5] == '' or values[key_c1] == '' or values[key_c2] == '':
                sg.Popup('Заполните все поля', title=errt, button_type=0, icon='warning.ico')
                break
            else:
                value = ['PULSE(', values[key_r1], pulse_power, values[key_r3], values[key_r4], values[key_r5],
                         values[key_c1],
                         values[key_c2], ')']
                res = ' '.join([str(item) for item in value])
                break

    window.close()
    return res


# In[17 ]

def power_sine(pulse_power, element_name):
    layout = [
        [sg.Text(text_enter_pulsentime + ' для элемента ' + element_name)],
        [sg.Text('DC offset[V]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r1), ],
        [sg.Text('Amplitude[V]: заполняется автоматически '), ],
        [sg.Text('Freq[Hz]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r3), ],
        [sg.Text('Tdelay[s]: ', size=(8, 0)), sg.InputText('1', size=(8, 5), key=key_r4), ],
        [sg.Text('Theta[1/s]: ', size=(8, 0)), sg.InputText('1', size=(8, 5), key=key_r5), ],
        [sg.Text('Phi[deg]: ', size=(8, 0)), sg.InputText('1000', size=(8, 5), key=key_c1), ],
        [sg.Text('Ncycles: ', size=(8, 0)), sg.InputText('10000', size=(8, 5), key=key_c2), ],
        [sg.Button('Ввести'), sg.Button('Cancel')],
        [sg.Text()], ]

    window = sg.Window('Ввод параметров симуляции', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            res = 'None'
            break
        elif event == 'Ввести':
            if values[key_r1] == '' or values[key_r3] == '' or values[key_r4] == '' or values[
                key_r5] == '' or values[key_c1] == '' or values[key_c2] == '':
                sg.Popup('Заполните все поля', title=errt, button_type=0, icon='warning.ico')
                break
            else:
                value = ['SINE(', values[key_r1], pulse_power, values[key_r3], values[key_r4], values[key_r5],
                         values[key_c1],
                         values[key_c2], ')']
                res = ' '.join([str(item) for item in value])
                break
    window.close()
    return res


# In[17 ]

def power_exp(pulse_power, element_name):
    layout = [
        [sg.Text(text_enter_pulsentime + ' для элемента ' + element_name)],
        [sg.Text('Vinitial[V]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r1), ],
        [sg.Text('Vpulsed[V]: заполняется автоматически '), ],
        [sg.Text('Rise Delay[s]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r3), ],
        [sg.Text('Rise Tau[s]: ', size=(8, 0)), sg.InputText('1', size=(8, 5), key=key_r4), ],
        [sg.Text('Fall Delay[s]: ', size=(8, 0)), sg.InputText('1', size=(8, 5), key=key_r5), ],
        [sg.Text('Fall Tau[s]: ', size=(8, 0)), sg.InputText('1000', size=(8, 5), key=key_c1), ],
        [sg.Button('Ввести'), sg.Button('Cancel')],
        [sg.Text()], ]

    window = sg.Window('Ввод параметров симуляции', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            res = 'None'
            break
        elif event == 'Ввести':
            if values[key_r1] == '' or values[key_r3] == '' or values[key_r4] == '' or values[
                key_r5] == '' or values[key_c1] == '':
                sg.Popup('Заполните все поля', title=errt, button_type=0, icon='warning.ico')
                break
            else:
                value = ['EXP(', values[key_r1], pulse_power, values[key_r3], values[key_r4], values[key_r5],
                         values[key_c1], ')']
                res = ' '.join([str(item) for item in value])
                break
    window.close()
    return res


# In[17 ]

def power_sffm(pulse_power, element_name):
    layout = [
        [sg.Text(text_enter_pulsentime + ' для элемента ' + element_name)],
        [sg.Text('DC offset[V]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r1), ],
        [sg.Text('Amplitude[V]: заполняется автоматически '), ],
        [sg.Text('Carrier Freq[Hz]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r3), ],
        [sg.Text('Modulation Index: ', size=(8, 0)), sg.InputText('1', size=(8, 5), key=key_r4), ],
        [sg.Text('Signal Freq[Hz]: ', size=(8, 0)), sg.InputText('1', size=(8, 5), key=key_r5), ],
        [sg.Button('Ввести'), sg.Button('Cancel')],
        [sg.Text()], ]

    window = sg.Window('Ввод параметров симуляции', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            res = 'None'
            break
        elif event == 'Ввести':
            if values[key_r1] == '' or values[key_r3] == '' or values[key_r4] == '' or values[
                key_r5] == '':
                sg.Popup('Заполните все поля', title=errt, button_type=0, icon='warning.ico')
                break
            else:
                value = ['SFFM(', values[key_r1], pulse_power, values[key_r3], values[key_r4], values[key_r5], ')']
                res = ' '.join([str(item) for item in value])
                break
    window.close()
    return res


# In[17 ]

def power_pwl(pulse_power, element_name):
    layout = [
        [sg.Text(text_enter_pulsentime + ' для элемента ' + element_name)],
        [sg.Text('time 1[s]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r1), ],
        [sg.Text('value 1[V]: ', size=(8, 0)), sg.InputText(pulse_power, size=(8, 5), key=key_r2), ],
        [sg.Text('time 2[s]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r3), ],
        [sg.Text('value 2[V]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r4), ],
        [sg.Text('time 3[s]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_r5), ],
        [sg.Text('value 3[V]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_c1), ],
        [sg.Text('time 4[s]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_c2), ],
        [sg.Text('value 4[V]: ', size=(8, 0)), sg.InputText('0', size=(8, 5), key=key_c3), ],
        [sg.Button('Ввести'), sg.Button('Cancel')],
        [sg.Text()], ]

    window = sg.Window('Ввод параметров симуляции', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            res = 'None'
            break
        elif event == 'Ввести':
            if values[key_r1] == '' or values[key_r2] == '' or values[key_r3] == '' or values[
                key_r4] == '' or values[key_r5] == '' or values[key_c1] == '' or values[key_c2] == '' or values[
                key_c3] == '':
                sg.Popup('Заполните все поля', title=errt, button_type=0, icon='warning.ico')
                break
            else:
                value = ['PWL(', values[key_r1], values[key_r2], values[key_r3], values[key_r4], values[key_r5],
                         values[key_c1],
                         values[key_c2], values[key_c3], ')']
                res = ' '.join([str(item) for item in value])
                break
    window.close()
    return res


# In[17 ]
def case_avgs(case_names):
    layout = [
        [sg.Text('Выберите корпус')],
        [sg.Listbox(case_names, s=(60, 10), select_mode='extended', key=key_list_rad)],
        [sg.Button('OK'), sg.Button('Cancel')],
        [sg.Text()],
    ]
    window = sg.Window('Усредненные значения корпусов', layout, finalize=True, modal=True)

    while True:
        event, values = window.read()
        window.find_element(key=key_list_rad).Update(case_names)
        if event in (sg.WINDOW_CLOSED, 'Cancel'):
            value = 'None'
            break
        elif event == 'OK':
            value = values[key_list_rad]
        if event == 'OK' and value != [] and len(value) == 1:
            break
    window.close()
    return value


# In[17 ]

def power_show(power, element_names):
    text = ''
    for i in range(len(power)):
        text = text + str(element_names[i]) + ' - ' + str(power[i]) + ' Вт\n'
    value = sg.Popup(text_power + '\n' + text, title='Мощности', custom_text=('OK', 'Cancel'), icon='icon.ico')
    return value


# In[17]

def ask_temp_change():
    ans = sg.Popup(text_ask_temp_change, title=title_temp_change, custom_text=('Да', 'Нет'), icon='icon.ico')
    return ans


# In[17]

def savebias_in_net(final_file, schema_name, tran_time):
    try:
        f = open(final_file, 'r+')
        line = f.readline()
        while line:
            a = f.tell()
            if line[0] == '.' and line[1] == 'e':
                f.seek(a - 4, 0)
                name = 'temperatures_' + schema_name.replace('.asc', '.txt')
                while ' ' in name:
                    name = name.replace(' ', '_')
                f.write('.savebias ' + name + ' time=' + tran_time + '\n.end')
                break
            line = f.readline()
        return name
    except FileNotFoundError:
        sg.Popup('Файл ' + final_file + ' не найден.', title=errt, button_type=0, icon='warning.ico')
        return 'None'


# In[17]

def read_savebias(svadrs, sv, name, net_format, temp_file_name=None):
    if net_format == 'cir':
        temp_file_name = temp_file_name.lower()
        full_adress = svadrs + '/' + temp_file_name
        mylix = file2strarr(full_adress)
        t = []
        if mylix != 'None':
            for line in mylix:
                gar = line.split()
                for z in gar:
                    if 'V(tj_' in z:
                        t.append('.PARAM ' + z.replace('V(tj', 'T').replace(')=', ' ').upper() + '\n')
            saving(t, sv, net_format)
    else:
        full_adress = svadrs + '/тепловая модель ' + name.replace('CKT', 'SVV')
        mylix = file2strarr(full_adress)
        t = []
        if mylix != 'None':
            for line in mylix:
                if 'TJ_' in line:
                    t.append('.DEFINE ' + line.replace('J', '').replace('    ', ' '))
            print('t', t)
            saving(t, sv, net_format)


# In[17]

def which_text_temp(net_format):
    if net_format == 'cir':
        return text_temp_change_spice
    else:
        return text_temp_change_microcap


# In[17]

def temp_change(name, svadrs, net_format, temp_file_name=None):
    text_temp_change = which_text_temp(net_format)
    sv = None
    layout = [
        [sg.Text(text_temp_change, size=(50, 0))],
        [sg.Button('Записать температуры')],  # sg.Button('Копировать значение темп.'),
        [sg.Text()],
    ]
    window = sg.Window(title_temp_change, layout, finalize=True, modal=True)
    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        elif event == 'Записать температуры':
            include_file = include_address(name, 2)
            sv = svadrs + '/' + include_file
            inmes = '.include ' + sv
            sg.clipboard_set(inmes)
            read_savebias(svadrs, sv, name, net_format, temp_file_name)
            break
    window.close()
    return sv


# In[179]

def get_index(name):
    name = str(get_column_letter(name))
    return name


# ## Wireframe
def netlist_creator(path, power, elements, element_names, final_file):
    path_to_file = path
    i = 0
    m = 'M'
    d = 'D'
    q = 'Q'
    x = 'X'
    f = open(final_file, 'w')
    f.write('Comment\n')
    Tj = 70
    Ta = 25

    for i in range(len(elements)):
        result = re.findall(m, elements[i])
        if len(result) > 0:
            list_number = 0
        else:
            result1 = re.findall(d, elements[i])
            if len(result1) > 0:
                list_number = 2
            else:
                result2 = re.findall(q, elements[i])
                if len(result2) > 0:
                    list_number = 3
                else:
                    result3 = re.findall(x, elements[i])
                    if len(result3) > 0:
                        list_number = 0
                    else:
                        list_number = 1

        wb = openpyxl.load_workbook(path_to_file)
        sheets_list = wb.sheetnames
        sheet_active = wb[sheets_list[list_number]]
        row_max = sheet_active.max_row

        search_text = element_names[i]
        print('Ищем:', search_text)

        row_min = 1
        component_checker = 0
        rad_checker = 0
        case_checker = 0
        self_checker = 0
        res_checker = 0
        nod_counter = 1
        nc = 1
        Rjc = 0
        data_from_net = 'None'

        row_min_min = row_min
        row_max_max = row_max
        while row_min_min <= row_max_max:
            row_min_min = str(row_min_min)

            word_cell = get_index(2) + row_min_min

            data_from_cell = str(sheet_active[word_cell].value)
            regular = search_text

            if (regular == data_from_cell) and (row_min_min != '1'):
                print('Нашли в ячейке:', word_cell)
                Net_list_column = 3
                Net_list_cell = get_index(Net_list_column) + row_min_min
                component_checker = 1
                column_net = Net_list_column

                while get_index(column_net) < 'M':
                    data_from_net = str(sheet_active[Net_list_cell].value)
                    Component_cell = get_index(column_net) + '1'
                    Component_name = str(sheet_active[Component_cell].value)

                    if data_from_net != '0.0' and data_from_net != '0':
                        if nod_counter % 2 != 0:
                            f.write(Component_name + '_' + elements[i] + ' ' + nods[nc + 1] + elements[i] + ' ' + nods[
                                nc] + elements[i] + ' ' + data_from_net + '\n')
                            Rjc += float(data_from_net)
                            nc += 1

                        if nod_counter % 2 == 0:
                            f.write(Component_name + '_' + elements[i] + ' ' + nods[
                                nod_counter // 2] + elements[i] + ' 0 ' + data_from_net + '\n')

                    nod_counter = nod_counter + 1
                    column_net = column_net + 1
                    Net_list_cell = get_index(column_net) + row_min_min

                data_from_net_ramb = str(sheet_active[Net_list_cell].value)

                column_net = column_net + 1
                Net_list_cell = get_index(column_net) + row_min_min
                data_from_net = str(sheet_active[Net_list_cell].value)

            row_min_min = int(row_min_min)
            row_min_min = row_min_min + 1

        rad_names = []

        if (row_min_min - 1) == row_max_max and component_checker == 0:
            print(regular + ' - Такого компонента нет в базе')
            if elements[i][0] == 'R' or elements[i][0] == 'r':
                comp_data_enter = self_enter_res(element_names[i])
                component_checker = 0
                res_checker = 1
                if comp_data_enter != 'None':
                    Rjc = write_comp(f, comp_data_enter, elements, i, nod_counter, nc, Rjc)
            else:
                ans = 'o'
                while ans != 'Выбрать усреднённые значения для корпуса' and ans != 'Ввести свои значения':
                    ans = sg.Popup(regular + ' - Такого компонента нет в базе', title='Тепловая модель',
                                   custom_text=('Выбрать усреднённые значения для корпуса', 'Ввести свои значения'),
                                   icon='icon.ico')
                    if ans == 'Выбрать усреднённые значения для корпуса':
                        component_checker = 1
                        case_names = []
                        sheet_active_case = wb[sheets_list[4]]
                        case_row = 2
                        case_name_cell = 'B' + str(case_row)
                        while str(sheet_active_case[case_name_cell].value) != 'None':
                            case_names.append(str(sheet_active_case[case_name_cell].value))
                            case_row += 1
                            case_name_cell = 'B' + str(case_row)

                        search_case = case_avgs(case_names)[0]  # .split()[0]

                        while str(sheet_active_case[case_name_cell].value) != 'Название корпусов':
                            if search_case == 'N':
                                ans = 'o'
                                break
                            else:
                                if search_case == str(sheet_active_case[case_name_cell].value):
                                    nod_counter = 0
                                    case_column = 2
                                    while get_index(case_column) < 'M':
                                        data_from_case = str(sheet_active_case[case_name_cell].value)
                                        Component_cell = get_index(case_column) + '1'
                                        Component_name = str(sheet_active_case[Component_cell].value)

                                        if data_from_case != '0.0' and data_from_case != '0' and data_from_case != search_case:
                                            if nod_counter % 2 != 0:
                                                f.write(
                                                    Component_name + '_' + elements[i] + ' ' + nods[nc + 1] + elements[
                                                        i] + ' ' + nods[
                                                        nc] + elements[i] + ' ' + data_from_case + '\n')
                                                Rjc += float(data_from_case)
                                                nc += 1

                                            if nod_counter % 2 == 0:
                                                f.write(Component_name + '_' + elements[i] + ' ' + nods[
                                                    nod_counter // 2] + elements[i] + ' 0 ' + data_from_case + '\n')

                                        nod_counter = nod_counter + 1
                                        case_column = case_column + 1
                                        case_name_cell = get_index(case_column) + str(case_row)

                                    data_from_net_ramb = str(sheet_active_case[case_name_cell].value)
                                    case_column = case_column + 1
                                    Net_list_cell = get_index(case_column) + str(case_row)
                                    data_from_net = str(sheet_active[Net_list_cell].value)
                                    case_checker = 1
                                    break
                                else:
                                    case_row -= 1
                                    case_name_cell = 'B' + str(case_row)

                    elif ans == 'Ввести свои значения':
                        comp_data_enter, r, ramb = self_enter(element_names[i])
                        if not r:
                            data_from_net = 'None'
                        else:
                            data_from_net = 'Possible'
                            # rad_checker = 1
                        if comp_data_enter != 'None':
                            component_checker = 1
                            self_checker = 1
                            Rjc = write_comp(f, comp_data_enter, elements, i, nod_counter, nc, Rjc)
                            data_from_net_ramb = ramb
                        else:
                            component_checker = 0
                        break

        answ = 'o'
        answr = 'o'
        if data_from_net != 'None' and rad_checker == 0:
            while answ != 'Да' and answ != 'Нет':
                answ = sg.Popup(text_rad + element_names[i] + '?', title=title_rad,
                                custom_text=('Да', 'Нет'), icon='icon.ico')
                if answ == 'Да':
                    while answr != 'Выбрать радиатор  из списка' and answr != 'Ввести свои значения':
                        answr = sg.Popup(text_choice, title=title_rad,
                                         custom_text=(
                                             'Выбрать радиатор  из списка', 'Ввести свои значения'),
                                         icon='icon.ico')
                        if answr == 'Выбрать радиатор  из списка':
                            if data_from_net == 'None':
                                sg.Popup('Для этого компонента нет радиатора', title=errt, button_type=0,
                                         icon='warning.ico')
                                # break
                            else:
                                R_rad = (Tj - Ta) / float(power[i]) - Rjc
                                sheet_active_rad = wb[sheets_list[1]]
                                row_max_rad = sheet_active_rad.max_row
                                column_max_rad = sheet_active_rad.max_column

                                for j in range(1, row_max_rad):
                                    j += 1
                                    word_cell_rads = 'B' + str(j)
                                    data_from_cell_rads = sheet_active_rad[word_cell_rads].value
                                    if str(data_from_cell_rads) != 'None':
                                        rad_name_cell = 'A' + str(j)
                                        rad_rth_cell = 'B' + str(j)
                                        rad_cth_cell = 'C' + str(j)
                                        rad_size_cell = 'D' + str(j)
                                        if R_rad < 2:
                                            if (R_rad) < float(data_from_cell_rads) < (
                                                    R_rad + 2):
                                                data_from_rad_name_cell = sheet_active_rad[
                                                                              rad_name_cell].value + ' (R_th = ' + \
                                                                          sheet_active_rad[
                                                                              rad_rth_cell].value \
                                                                          + ', C_th = ' + sheet_active_rad[
                                                                              rad_cth_cell].value + ', Size = ' + \
                                                                          sheet_active_rad[
                                                                              rad_size_cell].value + ')'
                                                rad_names.append(data_from_rad_name_cell)
                                        else:
                                            if (R_rad - R_rad * 0.2) < float(data_from_cell_rads) < (
                                                    R_rad + R_rad * 0.2):
                                                data_from_rad_name_cell = sheet_active_rad[
                                                                              rad_name_cell].value + ' (R_th = ' + \
                                                                          sheet_active_rad[
                                                                              rad_rth_cell].value \
                                                                          + ', C_th = ' + sheet_active_rad[
                                                                              rad_cth_cell].value + ', Size = ' + \
                                                                          sheet_active_rad[
                                                                              rad_size_cell].value + ')'
                                                rad_names.append(data_from_rad_name_cell)
                                if rad_names == []:
                                    sg.Popup('Для этого компонента нет радиатора', title=errt,
                                             button_type=0, icon='warning.ico')
                                    answr = 'Ввести свои значения'
                                else:
                                    search_text_rad = rad_name_pick(rad_names, element_names[i])[0].split()[0]
                                    row_min_rad = 1
                                    column_min_rad = 1

                                    while column_min_rad <= column_max_rad:
                                        if search_text_rad == 'N':
                                            answr = 'o'
                                            break
                                        else:
                                            row_min_min_rad = row_min_rad
                                            row_max_max_rad = row_max_rad
                                            while row_min_min_rad <= row_max_max_rad:
                                                row_min_min_rad = str(row_min_min_rad)

                                                word_cell_rad = get_index(column_min_rad) + row_min_min_rad

                                                data_from_cell_rad = str(
                                                    sheet_active_rad[word_cell_rad].value)
                                                regular_rad = search_text_rad

                                                if regular_rad == data_from_cell_rad:
                                                    print('Нашли в ячейке:', word_cell_rad)
                                                    Net_list_column_rad = 2
                                                    Net_list_cell_rad = get_index(
                                                        Net_list_column_rad) + row_min_min_rad
                                                    rad_checker = 1
                                                    column_net_rad = Net_list_column_rad

                                                    while get_index(column_net_rad) < 'D':
                                                        data_from_net_rad = str(
                                                            sheet_active_rad[Net_list_cell_rad].value)
                                                        Component_cell_rad = get_index(column_net_rad) + '1'
                                                        Component_name_rad = str(
                                                            sheet_active_rad[Component_cell_rad].value)

                                                        if column_net_rad == 2:
                                                            f.write(
                                                                Component_name_rad + '_' + elements[i] + ' ' +
                                                                nods[8] + ' ' + nods[
                                                                    nc] + elements[i] + ' ' + data_from_net_rad + '\n')

                                                        else:
                                                            f.write(
                                                                Component_name_rad + '_' + elements[i] + ' ' +
                                                                nods[
                                                                    nc] + elements[
                                                                    i] + ' 0 ' + data_from_net_rad + '\n')

                                                        column_net_rad = column_net_rad + 1
                                                        Net_list_cell_rad = get_index(
                                                            column_net_rad) + row_min_min_rad

                                                if int(row_min_min_rad) == row_max_max_rad and column_min_rad == column_max_rad and rad_checker == 0:
                                                    print(regular_rad + ' - Такого радиатора нет в базе')

                                                row_min_min_rad = int(row_min_min_rad)
                                                row_min_min_rad = row_min_min_rad + 1
                                            column_min_rad = column_min_rad + 1
                        elif answr == 'Ввести свои значения':
                            r_data_enter = self_enter_rad()
                            if r_data_enter != 'None':
                                rad_checker = 1
                                f.write('R_rad ' + nods[8] + ' ' + nods[nc] + elements[i] + ' ' + r_data_enter[
                                    0] + '\n')
                                f.write('C_rad ' + nods[nc] + elements[i] + ' 0 ' + r_data_enter[1] + '\n')
                            else:
                                answr = 'o'

                elif answ == 'Нет':
                    break

        if component_checker == 1 and rad_checker == 0:  # ) or answ == 'Нет' or answ == 'o'
            if data_from_net_ramb != '0.0' and data_from_net_ramb != '':
                if case_checker == 0 and self_checker == 0:
                    Component_cell = get_index(column_net - 1) + '1'
                    Component_name = str(sheet_active[Component_cell].value)
                    f.write(Component_name + '_' + elements[i] + ' Tamb ' + nods[nc] + elements[
                        i] + ' ' + data_from_net_ramb + '\n')
                elif case_checker == 1:
                    Component_cell = get_index(case_column - 1) + '1'
                    Component_name = str(sheet_active_case[Component_cell].value)
                    f.write(Component_name + '_' + elements[i] + ' Tamb ' + nods[nc] + elements[
                        i] + ' ' + data_from_net_ramb + '\n')
                elif self_checker == 1:
                    Component_name = 'R_case_to_ambient'
                    f.write(Component_name + '_' + elements[i] + ' Tamb ' + nods[nc] + elements[
                        i] + ' ' + data_from_net_ramb + '\n')

        if component_checker == 1 or res_checker == 1:
            power_value = choose_power(power[i], element_names[i])
            f.write('I_' + elements[i] + ' 0 Tj_' + elements[i] + ' ' + power_value + '\n')
    if i == len(element_names) - 1:
        t = enter_tran()
    f.write('Vamb Tamb 0 25\n' + '.tran ' + t + '\n' + '.end')
    f.close()
    return t


# In[170]:

def netlist_was_created(final_file):
    netlist_was_created = sg.Popup(final_popup_text, title='Готово', button_type=2,
                                   custom_text=open_net_file, icon='icon.ico')
    if netlist_was_created == open_net_file:
        try:
            os.startfile(final_file)
        except FileNotFoundError:
            sg.Popup('Файл ' + final_file + ' не найден.', title=errt, button_type=0, icon='warning.ico')


# In[170]:
layout = [
    [sg.Text(text_hint)],
    [sg.T("")], [sg.Text(text_browse_file), sg.Input(), sg.FileBrowse(button_text=text_open, button_color='green',
                                                                      file_types=(("Text Files", "*.txt"),
                                                                                  ("Text Files", "*.net"),
                                                                                  ("Text Files", "*.cir"),
                                                                                  ("Text Files", "*.CKT")
                                                                                  ),
                                                                      key=key_open)],
    [sg.T("")], [sg.Text(text_browse_excel), sg.Input(), sg.FileBrowse(button_text=text_open, button_color='green',
                                                                       file_types=(("Text Files", "*.xlsx"),),
                                                                       key=key_open_net)],
    [sg.Button(text_show_the_list, button_color='green')],
    [sg.Text(text_choose_elements), sg.Listbox(k, s=(25, 10), select_mode='extended', key=key_list)],
    [sg.Text(text_help)],
    [sg.Button(text_save, button_color='green')],
    [sg.Text(text_for_error_file_open)],
    [sg.Button(text_open_error_default, button_color='green')],
    [sg.Text(text_for_somebody)],
    [sg.Button(text_exit_button, button_color='red')]
]

# ## Using

# In[ ]:


# здесь работа программы самой sg. — функции библиотеки PySimpleGUI
# PopUp — всплывающие окошки, многочисленные else и elif нужны для вывода ошибок пользователя
# все переменные вроде key..что-нибудь или tetx..что-нибудь или errt или все остальное — тексты заданные в variables,
# для того, чтобы было удобней менять текст
# файлы формата .ico — мультииконки, они лежат в архиве с программой
window = sg.Window('SPICE Thermal Models Creator', layout, icon='icon.ico')

while True:
    event, values = window.read()

    if event in (sg.WIN_CLOSED, text_exit_button):
        window.close()
        break

    elif event == text_show_the_list:
        # values[<что-нибудь>] — зачения, полученные после какого-либо ивента,
        # в данном случае, функция открывает файл, но если открывать нечего, она выводит PopUp с ошибкой
        if values[key_open] == '':
            sg.Popup(chft, title=errt, button_type=0, icon='warning.ico')

        else:
            adrs = (values[key_open])
            net_format = 'cir'
            if '.CKT' in adrs:
                net_format = 'ckt'
            buf, k, schema_file, info_list = listreader(adrs)
            # print(buf) #- весь нетлист в виде массива из строк
            # print(k) - Названия элементов
            # print(schema_file) - адрес схемы
            # print(info_list) - строки нетлиста с элементами
            if k != None and schema_file != None:
                schema_address, schema_name = os.path.split(schema_file)
                window.find_element(key=key_list).Update(k)
    # условия типа event равно <что-нибудь>: <что-нибудь> это, в данном случае,
    # нажатие на кнопку с текстом в переменной "text_save"
    elif event == text_save:

        if values[key_open] == '':
            sg.Popup(chft, title=errt, button_type=0, icon='warning.ico')
        elif values[key_list] == []:
            sg.Popup(ches, title=errt2, button_type=0, icon='warning.ico')

        else:
            svadrs, name = os.path.split(adrs)
            include_file = include_address(name, 1)
            sv = svadrs + '/' + include_file  # - Адрес инклюда
            list_of_chosen_elements = values[key_list]
            first_name, number_of_elements = (list_of_chosen_elements[0].lower().replace('x', '') + '_power',
                                              len(list_of_chosen_elements))
            slist = []
            for val in values[key_list]:
                slist.append(val)

            if slist == []:
                sg.Popup(ches, title=errt2, button_type=0, icon='warning.ico')

            else:
                element_names = []
                chosen_elems_info = []
                path = (values[key_open_net])
                if info_list != None:
                    check = saving(info_list, adrs, net_format, slist, path)

                    for stroka in info_list:
                        stroka = stroka.split()
                        for elem in slist:
                            if elem in stroka:
                                elem_name, elem_stroka = give_name(stroka, net_format)
                                element_names.append(elem_name)
                                chosen_elems_info.append(elem_stroka)
                chosen_elems_info = replacing_lishnee(chosen_elems_info, element_names)
                if check:
                    spice_treatment = sg.Popup(message_treatment, title='Внимание', button_type=2,
                                               custom_text='Открыть файл', background_color='white', icon='icon.ico')

                    if spice_treatment == 'Открыть файл':
                        os.startfile(adrs)  # schema_file
                else:
                    sg.Popup('Не удалось записать формулы для рассчета мощностей.', title=errt2, button_type=0,
                             icon='warning.ico')

    elif event == text_open_error_default:

        if values[key_open] == '':
            sg.Popup(chft, title=errt, button_type=0, icon='warning.ico')
        elif values[key_open_net] == '':
            sg.Popup(chft2, title=errt, button_type=0, icon='warning.ico')
        else:
            original_error_file = error_file_adrs(adrs, net_format)
            original_error_address, original_error_name = os.path.split(original_error_file)
            directory_test_list = os.listdir(original_error_address)

            if original_error_name in directory_test_list:

                if slist == []:
                    sg.Popup(ches, title=errt2, button_type=0, icon='warning.ico')

                else:
                    final_file, final_file_w_temp = final_file_adrs_creator(svadrs, name)

                    power = power_file_saving(original_error_file, '_power', number_of_elements, net_format,
                                              list_of_chosen_elements)  # first_name

                    if (len(power) > 0) and (len(power) == len(slist)):
                        if element_names != []:
                            continue_program = power_show(power, element_names)
                            if continue_program == 'OK':
                                if new_net_w_temp == 0:
                                    tran_time = netlist_creator(path, power, slist, element_names, final_file)
                                else:
                                    tran_time = netlist_creator(path, power, slist, element_names, final_file_w_temp)
                                t = ask_temp_change()
                                if t == 'Да':
                                    if new_net_w_temp == 0:
                                        if '.cir' in final_file:
                                            temp_file_name = savebias_in_net(final_file, schema_name, tran_time)
                                            netlist_was_created(final_file)
                                        else:
                                            netlist_was_created(final_file)
                                    else:
                                        if '.cir' in final_file:
                                            temp_file_name = savebias_in_net(final_file_w_temp, schema_name, tran_time)
                                            netlist_was_created(final_file_w_temp)
                                        else:
                                            netlist_was_created(final_file_w_temp)
                                    if '.cir' in final_file:
                                        include_adrs = temp_change(name, svadrs, net_format, temp_file_name)
                                    else:
                                        include_adrs = temp_change(name, svadrs, net_format)
                                    if include_adrs != None:
                                        saving_temp(chosen_elems_info, adrs, net_format, include_adrs, slist)
                                        new_net_w_temp = 1
                                else:
                                    if new_net_w_temp == 1:
                                        netlist_was_created(final_file_w_temp)
                                    else:
                                        netlist_was_created(final_file)
                        else:
                            sg.Popup('Ошибка: не получены имена элементов.', title=errt, button_type=0,
                                     icon='warning.ico')
                    else:
                        sg.Popup('Ошибка: не посчитана мощность.', title=errt, button_type=0, icon='warning.ico')

            else:
                sg.Popup(error_text_error_log, title=errt, button_type=0, icon='warning.ico')
